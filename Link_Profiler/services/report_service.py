"""
Report Service - Manages the generation of various types of reports (e.g., PDF, Excel).
File: Link_Profiler/services/report_service.py
"""

import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
import io

# Placeholder for PDF generation library. ReportLab is a common choice.
# You might need to install it: pip install reportlab
# For more complex HTML-to-PDF, consider WeasyPrint (requires C libraries) or xhtml2pdf.
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    logging.warning("ReportLab not installed. PDF report generation will be simulated. Install with 'pip install reportlab'.")
    REPORTLAB_AVAILABLE = False

# Excel imports
# Declare names as None to satisfy flake8's F821 when imports are conditional
openpyxl = None
Font = None
PatternFill = None
get_column_letter = None
EXCEL_AVAILABLE = False # Initialize outside try block

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    logging.warning("openpyxl not installed. Excel export will be unavailable. Install with 'pip install openpyxl'.")
    # EXCEL_AVAILABLE remains False

from Link_Profiler.database.database import Database
from Link_Profiler.core.models import LinkProfile, Backlink, serialize_model
from Link_Profiler.utils.data_exporter import export_to_excel, EXCEL_AVAILABLE # New: Import export_to_excel and EXCEL_AVAILABLE

logger = logging.getLogger(__name__)

# Helper for ReportLab units (moved outside class for global access if needed)
inch = 72

class ReportService:
    """
    Service for generating various types of reports.
    """
    def __init__(self, database: Database):
        self.db = database
        self.logger = logging.getLogger(__name__)
        self.REPORTLAB_AVAILABLE = REPORTLAB_AVAILABLE # Expose availability for external check
        self.EXCEL_AVAILABLE = EXCEL_AVAILABLE # Expose availability for external check

    async def __aenter__(self):
        """No specific async setup needed for this class."""
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """No specific async cleanup needed for this class."""
        pass

    def _create_dummy_pdf(self, link_profile: LinkProfile, target_url: str, *, reason: str) -> io.BytesIO:
        """Return a simple dummy PDF representation as a placeholder."""
        dummy_pdf_content = (
            f"Simulated PDF Report for Link Profile: {target_url}\n\n"
            f"Total Backlinks: {link_profile.total_backlinks}\n"
            f"Unique Domains: {link_profile.unique_domains}\n"
            f"Authority Score: {link_profile.authority_score}\n"
            f"(Dummy PDF generated because: {reason})"
        )
        return io.BytesIO(dummy_pdf_content.encode("utf-8"))

    async def generate_link_profile_pdf_report(self, target_url: str) -> Optional[io.BytesIO]:
        """
        Generates a PDF report for a given link profile.
        Returns a BytesIO object containing the PDF data.
        """
        link_profile = self.db.get_link_profile(target_url)
        if not link_profile:
            self.logger.warning(f"Link profile not found for {target_url}. Cannot generate PDF report.")
            return None

        if not self.REPORTLAB_AVAILABLE:
            self.logger.warning(
                "ReportLab is not installed. Falling back to a dummy PDF report."
            )
            return self._create_dummy_pdf(link_profile, target_url, reason="ReportLab missing")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()

        story = []

        # Title
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['h1'],
            fontSize=24,
            leading=28,
            alignment=1, # Center
            spaceAfter=12
        )
        story.append(Paragraph("Link Profile Report", title_style))
        story.append(Paragraph(f"For: <font color='blue'>{link_profile.target_url}</font>", styles['h2']))
        story.append(Spacer(1, 0.2 * inch))

        # Summary Table
        summary_data = [
            ['Metric', 'Value'],
            ['Target Domain', link_profile.target_domain],
            ['Total Backlinks', str(link_profile.total_backlinks)],
            ['Unique Referring Domains', str(link_profile.unique_domains)],
            ['Dofollow Links', str(link_profile.dofollow_links)],
            ['Nofollow Links', str(link_profile.nofollow_links)],
            ['Authority Score', f"{link_profile.authority_score:.2f}"],
            ['Trust Score', f"{link_profile.trust_score:.2f}"],
            ['Spam Score', f"{link_profile.spam_score:.2f}"],
            ['Analysis Date', link_profile.analysis_date.strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        summary_table = Table(summary_data)
        summary_table.setStyle(table_style)
        story.append(summary_table)
        story.append(Spacer(1, 0.4 * inch))

        # Top Anchor Texts
        story.append(Paragraph("<h3>Top Anchor Texts</h3>", styles['h3']))
        if link_profile.anchor_text_distribution:
            # Sort anchor texts by count descending
            sorted_anchors = sorted(link_profile.anchor_text_distribution.items(), key=lambda item: item[1], reverse=True)[:10]
            anchor_data = [['Anchor Text', 'Count']]
            for text, count in sorted_anchors:
                anchor_data.append([text, str(count)])
            
            anchor_table = Table(anchor_data)
            anchor_table.setStyle(table_style) # Reuse table style
            story.append(anchor_table)
        else:
            story.append(Paragraph("No anchor text data available.", styles['Normal']))
        story.append(Spacer(1, 0.4 * inch))

        # Referring Domains (Top 10)
        story.append(Paragraph("<h3>Top Referring Domains</h3>", styles['h3']))
        if link_profile.referring_domains:
            # Convert set to list and take top 10 (alphabetical for consistency)
            sorted_domains = sorted(list(link_profile.referring_domains))[:10]
            domain_data = [['Domain']]
            for domain_name in sorted_domains:
                domain_data.append([domain_name])
            
            domain_table = Table(domain_data)
            domain_table.setStyle(table_style)
            story.append(domain_table)
        else:
            story.append(Paragraph("No referring domain data available.", styles['Normal']))
        story.append(Spacer(1, 0.4 * inch))

        # Backlinks (Sample)
        story.append(Paragraph("<h3>Sample Backlinks</h3>", styles['h3']))
        if link_profile.backlinks:
            backlink_header = ['Source URL', 'Anchor Text', 'Type', 'Spam Level']
            backlink_rows = []
            for bl in link_profile.backlinks[:10]: # Limit to first 10 for sample
                backlink_rows.append([
                    bl.source_url,
                    bl.anchor_text,
                    bl.link_type.value,
                    bl.spam_level.value
                ])
            
            backlink_table_data = [backlink_header] + backlink_rows
            backlink_table = Table(backlink_table_data, colWidths=[2.5*inch, 1.5*inch, 0.8*inch, 0.8*inch])
            backlink_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            story.append(backlink_table)
        else:
            story.append(Paragraph("No detailed backlink data available.", styles['Normal']))
        
        # Build PDF
        try:
            doc.build(story)
            buffer.seek(0)
            self.logger.info(f"PDF report generated for link profile: {target_url}")
            return buffer
        except Exception as e:
            self.logger.error(
                f"Error building PDF for {target_url}: {e}. Using dummy PDF instead.",
                exc_info=True,
            )
            return self._create_dummy_pdf(link_profile, target_url, reason="PDF generation error")

    async def generate_link_profile_excel_report(self, target_url: str) -> Optional[io.BytesIO]:
        """
        Generates an Excel report for a given link profile.
        Returns a BytesIO object containing the XLSX data.
        """
        link_profile = self.db.get_link_profile(target_url)
        if not link_profile:
            self.logger.warning(f"Link profile not found for {target_url}. Cannot generate Excel report.")
            return None

        if not self.EXCEL_AVAILABLE:
            self.logger.error("openpyxl is not installed. Cannot generate Excel report. Returning None.")
            return None

        # Prepare data for Excel export
        # Main Summary Sheet
        summary_data = [
            {"Metric": "Target URL", "Value": link_profile.target_url},
            {"Metric": "Target Domain", "Value": link_profile.target_domain},
            {"Metric": "Total Backlinks", "Value": link_profile.total_backlinks},
            {"Metric": "Unique Referring Domains", "Value": link_profile.unique_domains},
            {"Metric": "Dofollow Links", "Value": link_profile.dofollow_links},
            {"Metric": "Nofollow Links", "Value": link_profile.nofollow_links},
            {"Metric": "Authority Score", "Value": f"{link_profile.authority_score:.2f}"},
            {"Metric": "Trust Score", "Value": f"{link_profile.trust_score:.2f}"},
            {"Metric": "Spam Score", "Value": f"{link_profile.spam_score:.2f}"},
            {"Metric": "Analysis Date", "Value": link_profile.analysis_date.strftime("%Y-%m-%d %H:%M:%S")}
        ]

        # Backlinks Sheet
        backlinks_data = []
        if link_profile.backlinks:
            for bl in link_profile.backlinks:
                backlinks_data.append(serialize_model(bl)) # Use serialize_model to convert Backlink to dict

        # Create a new workbook and add sheets
        workbook = openpyxl.Workbook()

        # Add Summary Sheet
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        summary_sheet.append(["Metric", "Value"])
        for row in summary_data:
            summary_sheet.append([row["Metric"], row["Value"]])
        
        # Style summary header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in summary_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-size columns for summary
        for col in summary_sheet.columns:
            max_length = 0
            column = col[0].column # Get the column number
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            summary_sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

        # Add Backlinks Sheet
        if backlinks_data:
            backlinks_sheet = workbook.create_sheet("Backlinks")
            # Infer fieldnames from the first backlink data dictionary
            backlink_fieldnames = list(backlinks_data[0].keys()) if backlinks_data else []
            backlinks_sheet.append(backlink_fieldnames)

            # Style backlinks header
            for cell in backlinks_sheet[1]:
                cell.font = header_font
                cell.fill = header_fill

            for row_data in backlinks_data:
                row_values = [row_data.get(key, '') for key in backlink_fieldnames]
                backlinks_sheet.append(row_values)
            
            # Auto-size columns for backlinks
            for col in backlinks_sheet.columns:
                max_length = 0
                column = col[0].column # Get the column number
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                backlinks_sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        self.logger.info(f"Excel report generated for link profile: {target_url}")
        return buffer
