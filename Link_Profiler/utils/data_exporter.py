"""
Data Exporter - Provides utilities for exporting data to various formats.
File: Link_Profiler/utils/data_exporter.py
"""

import csv
import io
from typing import List, Dict, Any, Optional
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    logging.warning("openpyxl not installed. Excel export will be unavailable. Install with 'pip install openpyxl'.")
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)

async def export_to_csv(data: List[Dict[str, Any]], fieldnames: Optional[List[str]] = None) -> io.StringIO:
    """
    Exports a list of dictionaries to a CSV formatted StringIO object.

    Args:
        data: A list of dictionaries, where each dictionary represents a row.
        fieldnames: An optional list of column headers. If not provided,
                    it will be inferred from the keys of the first dictionary.

    Returns:
        A StringIO object containing the CSV data.
    """
    if not data:
        logger.warning("No data provided for CSV export.")
        output = io.StringIO()
        if fieldnames:
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
        return output

    if fieldnames is None:
        # Infer fieldnames from the keys of all dictionaries to ensure all columns are captured
        all_keys = set()
        for row in data:
            all_keys.update(row.keys())
        fieldnames = sorted(list(all_keys)) # Sort for consistent column order

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)

    writer.writeheader()
    for row in data:
        # Filter row to only include keys present in fieldnames, and handle missing keys
        filtered_row = {k: row.get(k, '') for k in fieldnames}
        writer.writerow(filtered_row)

    output.seek(0)
    logger.info(f"Exported {len(data)} records to CSV.")
    return output

async def export_to_excel(data: List[Dict[str, Any]], sheet_name: str = "Sheet1", fieldnames: Optional[List[str]] = None) -> Optional[io.BytesIO]:
    """
    Exports a list of dictionaries to an Excel (XLSX) formatted BytesIO object.

    Args:
        data: A list of dictionaries, where each dictionary represents a row.
        sheet_name: The name of the worksheet.
        fieldnames: An optional list of column headers. If not provided,
                    it will be inferred from the keys of the first dictionary.

    Returns:
        A BytesIO object containing the XLSX data, or None if openpyxl is not available.
    """
    if not EXCEL_AVAILABLE:
        logger.error("openpyxl is not installed. Cannot generate Excel report.")
        return None

    if not data:
        logger.warning("No data provided for Excel export.")
        # Create an empty workbook with headers if fieldnames are provided
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        if fieldnames:
            sheet.append(fieldnames)
            # Apply header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    if fieldnames is None:
        all_keys = set()
        for row in data:
            all_keys.update(row.keys())
        fieldnames = sorted(list(all_keys))

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    # Write header row
    sheet.append(fieldnames)

    # Apply header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Write data rows
    for row_data in data:
        row_values = [row_data.get(key, '') for key in fieldnames]
        sheet.append(row_values)

    # Auto-size columns
    for col in sheet.columns:
        max_length = 0
        column = col[0].column # Get the column number
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    logger.info(f"Exported {len(data)} records to Excel sheet '{sheet_name}'.")
    return buffer
